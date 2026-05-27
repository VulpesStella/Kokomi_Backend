import os
import logging
import pymysql
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from dotenv import load_dotenv

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

ROOT_DIR = Path(os.getcwd())
if (ROOT_DIR / 'env.dev').exists():
    load_dotenv('env.dev')
elif (ROOT_DIR / 'env.prod').exists():
    load_dotenv('env.prod')
else:
    raise FileNotFoundError('No environment file found')

DB_CONFIG = {
    "host": 'localhost',
    "port": int(os.getenv("MYSQL_PORT", 3306)),
    "user": 'root',
    "password": os.getenv("MYSQL_ROOT_PASSWORD"),
    "database": os.getenv("MYSQL_DATABASE"),
    "autocommit": True
}

class ShipDataExporter:
    """
    Universal ship data export framework.
    Initializes with database configuration and output file path.
    """

    def __init__(self, db_config: dict, output_path: Path):
        self.db_config = db_config
        self.output_path = output_path
        self.conn = pymysql.connect(**self.db_config)
        self.ships = []
        self._load_base_ships()

    def _load_base_ships(self):
        cursor = self.conn.cursor()
        try:
            sql = """
                SELECT
                    b.ship_id,
                    b.tier,
                    t.name AS type,
                    n.name AS nation,
                    a.zh_sg AS ship_name
                FROM T_ship_base b
                INNER JOIN T_ship_name a ON b.ship_id = a.ship_id
                INNER JOIN D_ship_type t ON b.type_id = t.id
                INNER JOIN D_ship_nation n ON b.nation_id = n.id
                WHERE b.is_enabled = 1 AND b.is_old = 0 
                ORDER BY b.ship_id
            """
            cursor.execute(sql)
            rows = cursor.fetchall()
            self.ships = []
            for row in rows:
                self.ships.append({
                    'ship_id': row[0],
                    'tier': row[1],
                    'type': row[2],
                    'nation': row[3],
                    'ship_name': row[4]
                })
            logger.info(f"Loaded {len(self.ships)} enabled ships")
        finally:
            cursor.close()

    def _apply_common_styling(self, ws, column_widths=None):
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        black_font = Font(name="SimHei")
        center_alignment = Alignment(horizontal='center', vertical='center')

        max_row = ws.max_row
        max_col = ws.max_column

        # 所有单元格居中对齐 + 黑体
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.alignment = center_alignment
                cell.font = black_font

        # 标题行边框
        for cell in ws[1]:
            cell.border = thin_border

        # 行高
        ws.row_dimensions[1].height = 25
        for row_idx in range(2, max_row + 1):
            ws.row_dimensions[row_idx].height = 13

        # 列宽：指定或自动调整
        if column_widths is not None:
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
        else:
            # 自动调整列宽：遍历每一列，计算内容最大长度
            for col_idx in range(1, max_col + 1):
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                max_len = 0
                for row_idx in range(1, max_row + 1):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value is not None:
                        # 转换为字符串并计算长度（中文字符按1个长度算，但可适当加宽）
                        cell_len = len(str(cell_value))
                        # 对于中文字符，可以乘以系数1.2，这里简单处理
                        max_len = max(max_len, cell_len)
                # 设置列宽，最多不超过50，最小为5
                adjusted_width = min(max_len + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width

        return ws

    def export_basic_info(self):
        if not self.ships:
            logger.warning("No ship data to export")
            return
        
        xlsx_title = 'ship_basic_info'

        output_path = self.output_path / f'{xlsx_title}.xlsx'
        output_path.parent.mkdir(parents=True, exist_ok=True)

        headers = ['ship_id', 'tier', 'type', 'nation', 'ship_name']
        titles = ['ShipID', 'Tier', 'Type', 'Nation', 'Name']

        wb = Workbook()
        ws = wb.active
        ws.title = xlsx_title
        ws.append(titles)

        for ship in self.ships:
            row = [ship.get(key, '') for key in headers]
            ws.append(row)

        col_widths = {
            'A':12,   # ship_id
            'B':6,    # tier
            'C':12,   # type
            'D':13,   # nation
            'E':25,   # ship_name
        }
        self._apply_common_styling(ws, col_widths)

        wb.save(output_path)
        logger.info(f"Exported basic info of {len(self.ships)} ships to {output_path}")

    def export_pvp_record(self):
        metric_ids = [3, 4, 5, 7, 8, 9]
        metric_names = ['Damage', 'Frags', 'Exp', 'ScoutingDmg', 'PotentialDmg', 'Planes']

        if not self.ships:
            logger.warning("No ship data to export")
            return

        xlsx_title = 'ship_pvp_record'
        output_path = self.output_path / f'{xlsx_title}.xlsx'
        output_path.parent.mkdir(parents=True, exist_ok=True)

        base_keys = ['ship_id', 'tier', 'type', 'nation', 'ship_name']
        base_titles = ['ShipID', 'Tier', 'Type', 'Nation', 'Name']
        titles = base_titles + metric_names

        wb = Workbook()
        ws = wb.active
        ws.title = xlsx_title
        ws.append(titles)

        cursor = self.conn.cursor()
        exported_count = 0
        skipped_count = 0

        try:
            for ship in self.ships:
                ship_id = ship['ship_id']
                metric_values = {name: None for name in metric_names}

                placeholders = ','.join(['%s'] * len(metric_ids))
                sql = f"""
                    SELECT metric_id, metric_value
                    FROM T_ship_pvp_record
                    WHERE ship_id = %s AND metric_id IN ({placeholders})
                """
                cursor.execute(sql, [ship_id] + metric_ids)
                rows = cursor.fetchall()

                damage_value = None
                for metric_id, value in rows:
                    if metric_id == 3:
                        damage_value = value
                        break
                # 如果 damage 为 0 或 None，跳过该船
                if damage_value is None or damage_value == 0:
                    skipped_count += 1
                    continue

                # 将查询结果映射到 metric_names
                for metric_id, value in rows:
                    idx = metric_ids.index(metric_id)
                    metric_values[metric_names[idx]] = value

                row = [ship.get(key, '') for key in base_keys]
                row += [metric_values.get(name, None) for name in metric_names]
                ws.append(row)
                exported_count += 1

        finally:
            cursor.close()

        if exported_count == 0:
            logger.warning(f"No ships exported (all skipped due to damage=0 or missing)")
        else:
            logger.info(f"Exported {exported_count} ships (skipped {skipped_count} ships with damage=0 or missing)")

        # 应用通用样式（自动调整列宽）

        col_widths = {
            'A':12,   # ship_id
            'B':6,    # tier
            'C':12,   # type
            'D':13,   # nation
            'E':25,   # ship_name
            'F':11,
            'G':8,
            'H':8,
            'I':13,
            'J':15,
            'K':8
        }
        self._apply_common_styling(ws, col_widths)

        wb.save(output_path)
        logger.info(f"PvP record file saved to {output_path}")

    def export_rating_distribution(self):
        if not self.ships:
            logger.warning("No ship data to export")
            return

        xlsx_title = 'ship_rating_distribution'
        output_path = self.output_path / f'{xlsx_title}.xlsx'
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # 基础列
        base_keys = ['ship_id', 'tier', 'type', 'nation', 'ship_name']
        base_titles = ['ShipID', 'Tier', 'Type', 'Nation', 'Name']
        # 分布列（排除 updated_at）
        dist_columns = [
            'sample_count', 'top1', 'top5', 'top10', 'top15',
            'top50', 'top75', 'top90'
        ]
        dist_titles = ['SampleCount', 'Top1%', 'Top5%', 'Top10%', 'Top15%',
                    'Top50%', 'Top75%', 'Top90%']

        headers = base_keys + dist_columns
        titles = base_titles + dist_titles

        wb = Workbook()
        ws = wb.active
        ws.title = xlsx_title
        ws.append(titles)

        # 批量查询所有启用船只的分布数据，且 sample_count >= 100
        ship_ids = [ship['ship_id'] for ship in self.ships]
        if not ship_ids:
            logger.warning("No valid ship IDs found")
            return

        placeholders = ','.join(['%s'] * len(ship_ids))
        sql = f"""
            SELECT 
                ship_id, sample_count, top1, top5, top10, top15,
                top50, top75, top90
            FROM T_ship_rating_distribution
            WHERE ship_id IN ({placeholders}) AND sample_count >= 100
        """
        cursor = self.conn.cursor()
        try:
            cursor.execute(sql, ship_ids)
            rows = cursor.fetchall()
            # 构建字典：ship_id -> distribution dict
            dist_map = {}
            for row in rows:
                dist_map[row[0]] = {
                    'sample_count': row[1],
                    'top1': row[2],
                    'top5': row[3],
                    'top10': row[4],
                    'top15': row[5],
                    'top50': row[6],
                    'top75': row[7],
                    'top90': row[8]
                }
        finally:
            cursor.close()

        # 填充数据行：只处理在 dist_map 中的船只（即符合 sample_count >= 100）
        exported_count = 0
        for ship in self.ships:
            ship_id = ship['ship_id']
            if ship_id not in dist_map:
                continue  # 跳过不符合条件的船只（无记录或 sample_count < 100）
            dist = dist_map[ship_id]
            row = [ship.get(key, '') for key in base_keys]
            row += [
                dist['sample_count'],
                dist['top1'],
                dist['top5'],
                dist['top10'],
                dist['top15'],
                dist['top50'],
                dist['top75'],
                dist['top90']
            ]
            ws.append(row)
            exported_count += 1

        if exported_count == 0:
            logger.warning("No ships found with sample_count >= 100")
            # 仍然保存一个仅含标题的空文件
            wb.save(output_path)
            logger.info(f"Empty file saved to {output_path}")
            return

        # 应用通用样式
        col_widths = {
            'A':12,   # ship_id
            'B':6,    # tier
            'C':12,   # type
            'D':13,   # nation
            'E':25,   # ship_name
            'F':13,
            'G':8,
            'H':8,
            'I':8,
            'J':8,
            'K':8,
            'L':8,
            'M':8
        }
        self._apply_common_styling(ws, col_widths)
        wb.save(output_path)
        logger.info(f"Exported rating distribution for {exported_count} ships (sample_count >= 100) to {output_path}")

    def close(self):
        if self.conn:
            self.conn.close()
            logger.info("Database connection closed")


def main():
    """Test the framework: export basic ship information to Excel."""
    output_file = ROOT_DIR / 'temp'
    exporter = None
    try:
        exporter = ShipDataExporter(DB_CONFIG, output_file)
        exporter.export_rating_distribution()
    finally:
        if exporter:
            exporter.close()


if __name__ == '__main__':
    # python tools\export_data.py
    try:
        main()
    except KeyboardInterrupt:
        logger.info("Interrupted by user")
    except Exception as e:
        logger.error(e)